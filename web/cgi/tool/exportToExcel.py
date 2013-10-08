#-*- coding:utf-8 -*  
import MySQLdb  
import time  
import sys  
from openpyxl.workbook import Workbook  
from openpyxl.writer.excel import ExcelWriter  
from openpyxl.cell import get_column_letter  

def export(title,header,rows):
    wb = Workbook()  
    ew = ExcelWriter(workbook = wb)  
    dest_filename = r'/opt/aoaola/web/html/excel/' + title +'.xlsx'  
    ws = wb.worksheets[0]  
    
    for x in range(1,len(header)+1):  
        col = get_column_letter(x)  
        ws.cell('%s%s'%(col, 1)).value = '%s' % (header[x-1])        

    for i in range(len(rows)):
        assert len(rows[i]) == 6
        for j in range(len(rows[i])):
            ws.cell(row=i+1,column=j).value = rows[i][j]

    ew.save(filename = dest_filename)  
